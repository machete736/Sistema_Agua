from decimal import Decimal
from functools import wraps
from datetime import date
from datetime import date, datetime

from django.db.models import Q, Sum, Count
from django.utils.dateparse import parse_date
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.contrib.auth import login, logout, update_session_auth_hash, get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.core.exceptions import ValidationError
from django.db.models import Q, Sum
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.db import transaction
from django.utils import timezone
from datetime import datetime
from .models import Socio, Medidor, Tarifa, Lectura, Cobro, Pago, QRGenerico, Afiliacion


import openpyxl
import re
import io
import requests
from PIL import Image
from decouple import config

Usuario = get_user_model()

MESES_ES = [
    'ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
    'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE'
]


def periodo_a_fecha(periodo):
    """
    Convierte '2026-06' a una fecha 2026-06-01.
    """
    return datetime.strptime(periodo + '-01', '%Y-%m-%d').date()


def periodo_nombre(periodo):
    """
    Convierte '2026-06' a 'JUNIO 2026'.
    """
    fecha = periodo_a_fecha(periodo)
    return f"{MESES_ES[fecha.month - 1]} {fecha.year}"


def siguiente_periodo(periodo):
    """
    Devuelve el siguiente periodo.
    Ej: '2026-06' -> '2026-07'
    Ej: '2026-12' -> '2027-01'
    """
    fecha = periodo_a_fecha(periodo)

    if fecha.month == 12:
        return f"{fecha.year + 1}-01"

    return f"{fecha.year}-{fecha.month + 1:02d}"


def obtener_periodos_faltantes(ultimo_periodo, nuevo_periodo):
    """
    Devuelve los meses faltantes entre la última lectura y el nuevo periodo.
    Ej: último 2026-06, nuevo 2026-09
    retorna ['2026-07', '2026-08']
    """
    faltantes = []

    actual = siguiente_periodo(ultimo_periodo)

    while periodo_a_fecha(actual) < periodo_a_fecha(nuevo_periodo):
        faltantes.append(actual)
        actual = siguiente_periodo(actual)

    return faltantes

# =============================================================
# PERMISOS POR ROL
# =============================================================

def rol_requerido(roles_permitidos):
    def decorador(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('login')

            if not hasattr(request.user, 'rol'):
                messages.error(request, 'Tu usuario no tiene un rol asignado.')
                return redirect('login')

            if request.user.rol == 'socio':
                messages.error(request, 'Los socios deben ingresar desde la aplicación móvil.')
                logout(request)
                return redirect('login')

            if request.user.rol not in roles_permitidos:
                messages.error(request, 'No tienes permiso para acceder a esta sección.')
                return redirect(redirect_por_rol(request.user))   # ← antes era 'dashboard' fijo

            return view_func(request, *args, **kwargs)
        return wrapper
    return decorador

def es_solo_admin(view_func):
    return rol_requerido(['ADMIN', 'admin'])(view_func)
def es_admin_o_tesorero(view_func):
    return rol_requerido(['ADMIN', 'admin', 'TESORERO', 'tesorero'])(view_func)


def es_admin_tesorero_o_lector(view_func):
    return rol_requerido(['ADMIN', 'admin', 'LECTOR', 'lector'])(view_func)


# =============================================================
# LOGIN / LOGOUT
# =============================================================
def redirect_por_rol(usuario):
    rol = getattr(usuario, 'rol', '').lower()
    if rol == 'lector':
        return 'lector_inicio'
    return 'dashboard'  # admin y tesorero
def login_view(request):
    if request.user.is_authenticated:
        if request.user.rol == 'socio':
            logout(request)
            messages.error(request, 'Los socios deben ingresar desde la aplicación móvil.')
            return redirect('login')
        return redirect(redirect_por_rol(request.user))

    form = AuthenticationForm(request, data=request.POST or None)

    if request.method == 'POST' and form.is_valid():
        usuario = form.get_user()

        if usuario.rol == 'socio':
            messages.error(request, 'Los socios deben ingresar desde la aplicación móvil.')
            return redirect('login')

        login(request, usuario)
        next_url = request.GET.get('next')
        return redirect(next_url or redirect_por_rol(usuario))

    return render(request, 'login.html', {'form': form})


@login_required
def logout_view(request):
    logout(request)
    return redirect('login')


# =============================================================
# CAMBIAR CONTRASEÑA
# =============================================================

@login_required
def cambiar_password_view(request):
    if request.user.rol == 'socio':
        messages.error(request, 'Los socios deben ingresar desde la aplicación móvil.')
        logout(request)
        return redirect('login')

    if request.method == 'POST':
        actual = request.POST.get('password_actual', '')
        nueva = request.POST.get('password_nuevo', '')
        confirmar = request.POST.get('password_confirmar', '')

        if not request.user.check_password(actual):
            messages.error(request, 'La contraseña actual no es correcta.')
        elif nueva != confirmar:
            messages.error(request, 'Las contraseñas nuevas no coinciden.')
        elif len(nueva) < 6:
            messages.error(request, 'La contraseña debe tener al menos 6 caracteres.')
        else:
            request.user.set_password(nueva)
            request.user.save()
            update_session_auth_hash(request, request.user)
            messages.success(request, 'Contraseña cambiada correctamente.')
            return redirect('dashboard')

    return render(request, 'cambiar_password.html')
# =============================================================
# REEMPLAZA tu función dashboard_view en views_web.py con esta
# =============================================================

@login_required
@es_admin_o_tesorero
def dashboard_view(request):
    hoy = date.today()
    periodo_actual = hoy.strftime('%Y-%m')

    # ── KPIs principales ──────────────────────────────────────
    total_socios = Socio.objects.filter(estado='ACTIVO').count()
    medidores_activos = Medidor.objects.filter(estado='Activo').count()

    estados_deuda = ['Pendiente', 'En Revision', 'En Revisión', 'Vencido']
    cobros_pendientes_qs = Cobro.objects.filter(estado_pago__in=estados_deuda)
    cobros_pendientes = cobros_pendientes_qs.count()

    cobrado_mes = Pago.objects.filter(
        fecha_pago__year=hoy.year,
        fecha_pago__month=hoy.month
    ).aggregate(total=Sum('monto_pagado'))['total'] or Decimal('0.00')

    deuda_total = cobros_pendientes_qs.aggregate(
        total=Sum('monto_total')
    )['total'] or Decimal('0.00')

    # ── Meta del mes: cuanto deberia cobrarse vs cuanto se cobro ──
    cobros_del_mes = Cobro.objects.filter(lectura__periodo=periodo_actual)
    meta_mes = cobros_del_mes.aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')
    porcentaje_meta = 0
    if meta_mes > 0:
        porcentaje_meta = round(float(cobrado_mes) / float(meta_mes) * 100)

    # ── Progreso de ronda de lecturas del mes actual ───────────
    medidores_con_lectura = Lectura.objects.filter(
        periodo=periodo_actual
    ).values('medidor').distinct().count()

    medidores_sin_lectura = medidores_activos - medidores_con_lectura
    porcentaje_lecturas = 0
    if medidores_activos > 0:
        porcentaje_lecturas = round(medidores_con_lectura / medidores_activos * 100)

    # ── Top 5 socios con mayor deuda acumulada ─────────────────
    top_morosos = Socio.objects.filter(
        estado='ACTIVO',
        recibos__estado_pago__in=estados_deuda
    ).annotate(
        deuda=Sum('recibos__monto_total')
    ).filter(deuda__gt=0).order_by('-deuda')[:5]

    # ── Cobros pendientes recientes (para la tabla) ────────────
    ultimos_pendientes = cobros_pendientes_qs.select_related(
        'socio', 'lectura'
    ).order_by('-fecha_emision')[:6]

    # ── Historial de cobros: ultimos 6 meses ──────────────────
    meses_cortos = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                     'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
    historial_labels = []
    historial_cobrado = []
    anio_actual, mes_actual = hoy.year, hoy.month
    meses_a_mostrar = []
    for i in range(5, -1, -1):
        m = mes_actual - i
        a = anio_actual
        while m <= 0:
            m += 12
            a -= 1
        meses_a_mostrar.append((a, m))

    for (anio, mes) in meses_a_mostrar:
        total_mes = Pago.objects.filter(
            fecha_pago__year=anio, fecha_pago__month=mes
        ).aggregate(total=Sum('monto_pagado'))['total'] or Decimal('0.00')
        historial_labels.append(f"{meses_cortos[mes - 1]} {anio}")
        historial_cobrado.append(float(total_mes))

    context = {
        'total_socios': total_socios,
        'medidores_activos': medidores_activos,
        'recibos_pendientes': cobros_pendientes,
        'cobrado_mes': cobrado_mes,
        'deuda_total': deuda_total,
        'meta_mes': meta_mes,
        'porcentaje_meta': porcentaje_meta,

        'medidores_con_lectura': medidores_con_lectura,
        'medidores_sin_lectura': medidores_sin_lectura,
        'porcentaje_lecturas': porcentaje_lecturas,
        'periodo_actual_legible': periodo_nombre(periodo_actual) if 'periodo_nombre' in dir() else periodo_actual,

        'top_morosos': top_morosos,
        'ultimos_pendientes': ultimos_pendientes,

        'historial_labels': historial_labels,
        'historial_cobrado': historial_cobrado,
    }
    return render(request, 'dashboard.html', context)
# =============================================================
# SOCIOS — reemplaza todas las funciones socio_* en views_web.py
# =============================================================

@login_required
@es_admin_o_tesorero
def socios_lista(request):
    q = request.GET.get('q', '').strip()
    estado = request.GET.get('estado', 'ACTIVO')  # por defecto solo activos

    socios = Socio.objects.all()

    # Filtro por estado
    if estado and estado != 'TODOS':
        socios = socios.filter(estado=estado)

    # Búsqueda por texto
    if q:
        socios = socios.filter(
            Q(nombre_completo__icontains=q) |
            Q(ci__icontains=q) |
            Q(codigo_cliente__icontains=q) |
            Q(telefono__icontains=q)
        )

    return render(request, 'socios/lista.html', {
        'socios': socios,
        'q': q,
        'estado': estado,
        'total': socios.count(),
    })


@login_required
@es_solo_admin
def socio_crear(request):
    if request.method == 'POST':
        ci = request.POST.get('ci', '').strip()
        nombre = request.POST.get('nombre_completo', '').strip()
        codigo = request.POST.get('codigo_cliente', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        estado = request.POST.get('estado', 'ACTIVO').strip()
        observacion_retiro = request.POST.get('observacion_retiro', '').strip()
        cobrar_afiliacion = request.POST.get('cobrar_afiliacion') == 'on'
        fecha_afiliacion = request.POST.get('fecha_afiliacion', '').strip()

        if not ci or not nombre:
            messages.error(request, 'CI y nombre completo son obligatorios.')
        elif Socio.objects.filter(ci=ci).exists():
            messages.error(request, f'Ya existe un socio con CI {ci}.')
        else:
            from datetime import date
            socio = Socio.objects.create(
                ci=ci,
                nombre_completo=nombre,
                codigo_cliente=codigo or None,
                telefono=telefono or None,
                estado=estado,
                observacion_retiro=observacion_retiro or None,
                fecha_retiro=date.today() if estado == 'RETIRADO' else None,
            )

            if cobrar_afiliacion:
                tarifa = Tarifa.objects.filter(activa=True).order_by('-id_tarifa').first()
                Afiliacion.objects.create(
                    socio=socio,
                    tipo='NUEVO',
                    monto=tarifa.costo_afiliacion if tarifa else Decimal('0.00'),
                    fecha_pago=parse_date(fecha_afiliacion) or date.today(),
                    registrado_por=request.user,
                )
                messages.success(request, f'Socio {nombre} creado correctamente, con afiliación registrada.')
            else:
                messages.success(request, f'Socio {nombre} creado correctamente.')

            return redirect('socios_lista')

    return render(request, 'socios/form.html', {'accion': 'Crear'})


@login_required
@es_solo_admin
def socio_editar(request, pk):
    socio = get_object_or_404(Socio, pk=pk)

    if request.method == 'POST':
        ci = request.POST.get('ci', '').strip()
        nombre = request.POST.get('nombre_completo', '').strip()
        codigo = request.POST.get('codigo_cliente', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        estado = request.POST.get('estado', 'ACTIVO').strip()
        observacion_retiro = request.POST.get('observacion_retiro', '').strip()

        if not ci or not nombre:
            messages.error(request, 'CI y nombre completo son obligatorios.')
        elif Socio.objects.filter(ci=ci).exclude(pk=socio.pk).exists():
            messages.error(request, f'Ya existe otro socio con CI {ci}.')
        else:
            from datetime import date
            socio.ci = ci
            socio.nombre_completo = nombre
            socio.codigo_cliente = codigo or None
            socio.telefono = telefono or None
            socio.estado = estado
            socio.observacion_retiro = observacion_retiro or None

            if estado == 'RETIRADO':
                if not socio.fecha_retiro:
                    socio.fecha_retiro = date.today()
            else:
                socio.fecha_retiro = None

            socio.save()
            messages.success(request, 'Socio actualizado correctamente.')
            return redirect('socios_lista')

    return render(request, 'socios/form.html', {
        'accion': 'Editar',
        'socio': socio,
    })


@login_required
@es_admin_o_tesorero
def socio_detalle(request, pk):
    socio = get_object_or_404(Socio, pk=pk)
    medidores = Medidor.objects.filter(
        Q(socio=socio) | Q(co_titulares=socio)
    ).distinct()
    
    # Aquí estaba el error, cambiamos Recibo por Cobro
    recibos = Cobro.objects.filter(socio=socio).order_by('-fecha_emision') 
    
    estados_deuda = ['Pendiente', 'En Revision', 'En Revisión', 'Vencido']
    deuda_total = recibos.filter(
        estado_pago__in=estados_deuda
    ).aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')

    return render(request, 'socios/detalle.html', {
        'socio': socio,
        'medidores': medidores,
        'recibos': recibos, # Mantenemos la variable en minúscula para no romper el HTML
        'deuda_total': deuda_total,
    })
@login_required
@es_admin_o_tesorero
def socio_eliminar(request, pk):
    """Marca el socio como RETIRADO — nunca elimina el registro."""
    socio = get_object_or_404(Socio, pk=pk)
    if request.method == 'POST':
        from datetime import date
        socio.estado = 'RETIRADO'
        if not socio.fecha_retiro:
            socio.fecha_retiro = date.today()
        socio.save()
        messages.success(request, f'Socio {socio.nombre_completo} marcado como retirado.')
        return redirect('socios_lista')
    return render(request, 'socios/confirmar_eliminar.html', {'socio': socio})
@login_required
@es_solo_admin
def socio_crear_usuario_movil(request, pk):
    socio = get_object_or_404(Socio, pk=pk)

    if socio.usuario:
        messages.warning(request, 'Este socio ya tiene un usuario móvil asociado.')
        return redirect('socio_detalle', pk=socio.pk)

    if not socio.ci:
        messages.error(request, 'El socio no tiene CI registrado. No se puede crear usuario móvil.')
        return redirect('socio_detalle', pk=socio.pk)

    username = socio.ci.strip()

    if Usuario.objects.filter(username=username).exists():
        messages.error(
            request,
            f'Ya existe un usuario con el nombre "{username}". Revisa usuarios antes de continuar.'
        )
        return redirect('socio_detalle', pk=socio.pk)

    if hasattr(Usuario, 'ci') and Usuario.objects.filter(ci=socio.ci).exists():
        messages.error(
            request,
            f'Ya existe un usuario registrado con el CI {socio.ci}.'
        )
        return redirect('socio_detalle', pk=socio.pk)

    if request.method == 'POST':
        usuario = Usuario.objects.create_user(
            username=username,
            password=username,
            first_name=socio.nombre_completo,
        )

        usuario.rol = 'socio'

        if hasattr(usuario, 'ci'):
            usuario.ci = socio.ci

        if hasattr(usuario, 'activo'):
            usuario.activo = True

        usuario.is_active = True
        usuario.save()

        socio.usuario = usuario
        socio.save()

        messages.success(
            request,
            f'Usuario móvil creado correctamente. Usuario: {username} | Contraseña temporal: {username}'
        )
        return redirect('socio_detalle', pk=socio.pk)

    return render(request, 'socios/crear_usuario_movil.html', {
        'socio': socio,
        'username': username,
    })
@login_required
@es_admin_o_tesorero
def socio_estado_cuenta(request, pk):
    from datetime import date
    socio = get_object_or_404(Socio, pk=pk)
    anio = request.GET.get('anio', date.today().year)

    # Medidores del socio
    medidores = Medidor.objects.filter(
        Q(socio=socio) | Q(co_titulares=socio)
    ).distinct()

    # Cobros de la gestión seleccionada
    recibos = Cobro.objects.select_related('lectura', 'lectura__medidor').filter(
        socio=socio,
        lectura__periodo__startswith=str(anio)
    ).order_by('lectura__periodo')

    # Pagos realizados dentro de la misma gestión
    pagos = Pago.objects.filter(
        recibo__socio=socio,
        recibo__lectura__periodo__startswith=str(anio)
    )

    # Cálculos y agregaciones
    consumo_total = Decimal('0.00')
    total_emitido = Decimal('0.00')
    total_retrasos = Decimal('0.00')

    for recibo in recibos:
        recibo.aplicar_recargo_automatico()
        consumo_total += recibo.lectura.consumo_cubos
        total_emitido += recibo.monto_total
        total_retrasos += recibo.recargo_falta_pago

    total_pagado = pagos.aggregate(t=Sum('monto_pagado'))['t'] or Decimal('0.00')
    total_pendiente = total_emitido - total_pagado
    if total_pendiente < Decimal('0.00'):
        total_pendiente = Decimal('0.00')

    return render(request, 'socios/estado_cuenta.html', {
        'socio': socio,
        'medidores': medidores,
        'recibos': recibos,
        'anio': anio,
        'consumo_total': consumo_total,
        'total_emitido': total_emitido,
        'total_pagado': total_pagado,
        'total_retrasos': total_retrasos,
        'total_pendiente': total_pendiente,
        'usuario_emisor': request.user,  # Pasa el usuario autenticado para la firma
    })
# =============================================================
# MEDIDORES
# LISTA Y DETALLE: ADMIN / TESORERO / LECTOR
# CREAR, EDITAR, ELIMINAR: SOLO ADMIN / TESORERO
# =============================================================

@login_required
@es_admin_tesorero_o_lector
def medidores_lista(request):
    q = request.GET.get('q', '').strip()

    medidores = Medidor.objects.select_related('socio').prefetch_related('co_titulares').all()

    if q:
        medidores = medidores.filter(
            Q(numero_medidor__icontains=q) |
            Q(socio__nombre_completo__icontains=q) |
            Q(socio__ci__icontains=q) |
            Q(co_titulares__nombre_completo__icontains=q) |
            Q(co_titulares__ci__icontains=q) |
            Q(manzano__icontains=q) |
            Q(parcela__icontains=q)
        ).distinct()

    return render(request, 'medidores/lista.html', {
        'medidores': medidores,
        'q': q,
    })


@login_required
@es_admin_tesorero_o_lector
def medidor_detalle(request, pk):
    medidor = get_object_or_404(
        Medidor.objects.select_related('socio').prefetch_related('co_titulares'),
        pk=pk
    )

    lecturas = Lectura.objects.filter(medidor=medidor).order_by('-fecha_lectura')
    recibos = Cobro.objects.filter(lectura__medidor=medidor).order_by('-fecha_emision')

    return render(request, 'medidores/detalle.html', {
        'medidor': medidor,
        'lecturas': lecturas,
        'recibos': recibos,
    })


@login_required
@es_solo_admin
def medidor_crear(request):
    socios = Socio.objects.all()

    if request.method == 'POST':
        socio_id = request.POST.get('socio')
        numero_medidor = request.POST.get('numero_medidor', '').strip()
        manzano = request.POST.get('manzano', '').strip()
        parcela = request.POST.get('parcela', '').strip()
        estado = request.POST.get('estado', 'Activo')
        co_titulares_ids = request.POST.getlist('co_titulares')

        socio = get_object_or_404(Socio, pk=socio_id)

        if numero_medidor and Medidor.objects.filter(numero_medidor=numero_medidor).exists():
            messages.error(request, f'Ya existe un medidor con número {numero_medidor}.')
        else:
            medidor = Medidor.objects.create(
                socio=socio,
                numero_medidor=numero_medidor or None,
                manzano=manzano or None,
                parcela=parcela or None,
                estado=estado,
            )

            if co_titulares_ids:
                medidor.co_titulares.set(co_titulares_ids)

            messages.success(request, 'Medidor creado correctamente.')
            return redirect('medidores_lista')

    return render(request, 'medidores/form.html', {
        'accion': 'Crear',
        'socios': socios,
    })


@login_required
@es_solo_admin
def medidor_editar(request, pk):
    medidor = get_object_or_404(Medidor, pk=pk)
    socios = Socio.objects.all()

    if request.method == 'POST':
        socio_id = request.POST.get('socio')
        numero_medidor = request.POST.get('numero_medidor', '').strip()
        manzano = request.POST.get('manzano', '').strip()
        parcela = request.POST.get('parcela', '').strip()
        estado = request.POST.get('estado', 'Activo')
        co_titulares_ids = request.POST.getlist('co_titulares')
        cobrar_afiliacion = request.POST.get('cobrar_afiliacion') == 'on'
        fecha_afiliacion = request.POST.get('fecha_afiliacion', '').strip()

        if numero_medidor and Medidor.objects.filter(numero_medidor=numero_medidor).exclude(pk=medidor.pk).exists():
            messages.error(request, f'Ya existe otro medidor con número {numero_medidor}.')
        else:
            nuevo_socio = get_object_or_404(Socio, pk=socio_id)
            cambio_titular = medidor.socio_id != nuevo_socio.pk

            medidor.socio = nuevo_socio
            medidor.numero_medidor = numero_medidor or None
            medidor.manzano = manzano or None
            medidor.parcela = parcela or None
            medidor.estado = estado
            medidor.save()
            medidor.co_titulares.set(co_titulares_ids)

            if cambio_titular and cobrar_afiliacion:
                if hasattr(nuevo_socio, 'afiliacion'):
                    messages.warning(
                        request,
                        f'{nuevo_socio.nombre_completo} ya tiene una afiliación registrada; '
                        'no se creó una nueva.'
                    )
                else:
                    tarifa = Tarifa.objects.filter(activa=True).order_by('-id_tarifa').first()
                    Afiliacion.objects.create(
                        socio=nuevo_socio,
                        medidor=medidor,
                        tipo='TRANSFERENCIA',
                        monto=tarifa.costo_afiliacion if tarifa else Decimal('0.00'),
                        fecha_pago=parse_date(fecha_afiliacion) or date.today(),
                        registrado_por=request.user,
                    )
                    messages.success(request, 'Medidor actualizado y afiliación de transferencia registrada.')
                    return redirect('medidores_lista')

            messages.success(request, 'Medidor actualizado correctamente.')
            return redirect('medidores_lista')

    return render(request, 'medidores/form.html', {
        'accion': 'Editar',
        'medidor': medidor,
        'socios': socios,
    })


@login_required
@es_admin_o_tesorero
def medidor_eliminar(request, pk):
    medidor = get_object_or_404(Medidor, pk=pk)

    if request.method == 'POST':
        medidor.delete()
        messages.success(request, 'Medidor eliminado correctamente.')
        return redirect('medidores_lista')

    return render(request, 'medidores/confirmar_eliminar.html', {'medidor': medidor})

# =============================================================
# LECTURAS
#AJAX
# =============================================================
@login_required
@es_admin_tesorero_o_lector
def ajax_datos_medidor(request, pk):
    medidor = get_object_or_404(
        Medidor.objects.select_related('socio'),
        pk=pk
    )

    ultima_lectura = Lectura.objects.filter(
        medidor=medidor
    ).order_by('-fecha_lectura').first()

    lectura_anterior = '0.00'
    ultimo_periodo = ''

    if ultima_lectura:
        lectura_anterior = str(ultima_lectura.lectura_actual)
        ultimo_periodo = ultima_lectura.periodo

    data = {
        'id_medidor': str(medidor.pk),
        'numero_medidor': medidor.numero_medidor or '',
        'socio_id': str(medidor.socio.pk),
        'socio_nombre': medidor.socio.nombre_completo,
        'socio_ci': medidor.socio.ci,
        'codigo_cliente': medidor.socio.codigo_cliente or '',
        'manzano': medidor.manzano or '',
        'parcela': medidor.parcela or '',
        'estado': medidor.estado,
        'lectura_anterior': lectura_anterior,
        'ultimo_periodo': ultimo_periodo,
    }

    return JsonResponse(data)
# =============================================================
# LECTURAS
# LISTA, CREAR, DETALLE: ADMIN / TESORERO / LECTOR
# EDITAR, ELIMINAR: SOLO ADMIN / TESORERO
# =============================================================

@login_required
@es_admin_tesorero_o_lector
def lecturas_lista(request):
    q = request.GET.get('q', '').strip()

    lecturas = Lectura.objects.select_related('medidor', 'medidor__socio').all()

    if q:
        lecturas = lecturas.filter(
            Q(periodo__icontains=q) |
            Q(medidor__numero_medidor__icontains=q) |
            Q(medidor__socio__nombre_completo__icontains=q) |
            Q(medidor__socio__ci__icontains=q)
        )

    return render(request, 'lecturas/lista.html', {
        'lecturas': lecturas,
        'q': q,
    })
@login_required
@es_admin_tesorero_o_lector
def lectura_crear(request):
    socios = Socio.objects.all().order_by('nombre_completo')
    medidores = Medidor.objects.select_related('socio').filter(
        estado='Activo'
    ).order_by('socio__nombre_completo', 'numero_medidor')

    if request.method == 'POST':
        medidor_id = request.POST.get('medidor')
        periodo = request.POST.get('periodo', '').strip()
        lectura_anterior = request.POST.get('lectura_anterior', '0')
        lectura_actual = request.POST.get('lectura_actual', '0')
        observacion = request.POST.get('observacion', '').strip()
        foto = request.FILES.get('foto_evidencia')

        if not medidor_id:
            messages.error(request, 'Debe seleccionar un medidor.')
            return redirect('lectura_crear')

        if not periodo:
            messages.error(request, 'Debe seleccionar el periodo.')
            return redirect('lectura_crear')

        medidor = get_object_or_404(Medidor, pk=medidor_id)

        # 1. Evitar lectura duplicada en el mismo mes
        if Lectura.objects.filter(medidor=medidor, periodo=periodo).exists():
            messages.error(
                request,
                f'Ya existe una lectura registrada para este medidor en {periodo_nombre(periodo)}.'
            )
            return redirect('lectura_crear')

        # 2. Verificar si se está saltando meses
        ultima_lectura = Lectura.objects.filter(
            medidor=medidor
        ).order_by('-periodo').first()

        if ultima_lectura:
            ultimo_periodo = ultima_lectura.periodo
            fecha_ultimo = periodo_a_fecha(ultimo_periodo)
            fecha_nuevo = periodo_a_fecha(periodo)

            if fecha_nuevo <= fecha_ultimo:
                messages.error(
                    request,
                    f'No se puede registrar {periodo_nombre(periodo)} porque la última lectura registrada '
                    f'de este medidor corresponde a {periodo_nombre(ultimo_periodo)}.'
                )
                return redirect('lectura_crear')

            periodo_esperado = siguiente_periodo(ultimo_periodo)

            if periodo != periodo_esperado:
                faltantes = obtener_periodos_faltantes(ultimo_periodo, periodo)
                faltantes_texto = ', '.join([periodo_nombre(p) for p in faltantes])

                messages.error(
                    request,
                    f'No puedes registrar {periodo_nombre(periodo)} todavía. '
                    f'Primero falta registrar: {faltantes_texto}.'
                )
                return redirect('lectura_crear')

        try:
            lectura_anterior_decimal = Decimal(lectura_anterior)
            lectura_actual_decimal = Decimal(lectura_actual)

            if lectura_actual_decimal < lectura_anterior_decimal:
                messages.error(request, 'La lectura actual no puede ser menor que la lectura anterior.')
                return redirect('lectura_crear')

            lectura = Lectura(
                medidor=medidor,
                periodo=periodo,
                lectura_anterior=lectura_anterior_decimal,
                lectura_actual=lectura_actual_decimal,
                creado_por=request.user,
            )

            if hasattr(lectura, 'observacion'):
                lectura.observacion = observacion

            if foto:
                lectura.foto_evidencia = foto

            lectura.full_clean()
            lectura.save()

            messages.success(
                request,
                f'Lectura de {periodo_nombre(periodo)} registrada correctamente. El cobro se generó automáticamente.'
            )
            return redirect('lecturas_lista')

        except Exception as e:
            messages.error(request, f'Error al registrar la lectura: {e}')

    return render(request, 'lecturas/form.html', {
        'accion': 'Crear',
        'socios': socios,
        'medidores': medidores,
    })
@login_required
@es_admin_tesorero_o_lector
def lectura_detalle(request, pk):
    lectura = get_object_or_404(
        Lectura.objects.select_related('medidor', 'medidor__socio'),
        pk=pk
    )

    recibo = getattr(lectura, 'recibo', None)

    return render(request, 'lecturas/detalle.html', {
        'lectura': lectura,
        'recibo': recibo,
    })


@login_required
@es_admin_o_tesorero
def lectura_editar(request, pk):
    lectura = get_object_or_404(Lectura, pk=pk)
    medidores = Medidor.objects.select_related('socio').all()

    if request.method == 'POST':
        medidor_id = request.POST.get('medidor')
        periodo = request.POST.get('periodo', '').strip()
        fecha_lectura = request.POST.get('fecha_lectura')
        lectura_anterior = request.POST.get('lectura_anterior', '0')
        lectura_actual = request.POST.get('lectura_actual', '0')
        foto = request.FILES.get('foto_evidencia')

        if Lectura.objects.filter(medidor_id=medidor_id, periodo=periodo).exclude(pk=lectura.pk).exists():
            messages.error(request, 'Ya existe otra lectura para este medidor en ese periodo.')
        else:
            try:
                lectura.medidor = get_object_or_404(Medidor, pk=medidor_id)
                lectura.periodo = periodo
                if fecha_lectura:
                    lectura.fecha_lectura = fecha_lectura
                lectura.lectura_anterior = Decimal(lectura_anterior)
                lectura.lectura_actual = Decimal(lectura_actual)

                if foto:
                    lectura.foto_evidencia = foto

                lectura.full_clean()
                lectura.save()

                messages.success(request, 'Lectura actualizada correctamente.')
                return redirect('lecturas_lista')

            except (ValidationError, ValueError) as e:
                messages.error(request, f'Error al actualizar la lectura: {e}')

    return render(request, 'lecturas/form.html', {
        'accion': 'Editar',
        'lectura': lectura,
        'medidores': medidores,
    })


@login_required
@es_admin_o_tesorero
def lectura_eliminar(request, pk):
    lectura = get_object_or_404(Lectura, pk=pk)

    if request.method == 'POST':
        lectura.delete()
        messages.success(request, 'Lectura eliminada correctamente.')
        return redirect('lecturas_lista')

    return render(request, 'lecturas/confirmar_eliminar.html', {'lectura': lectura})
def llamar_ocr_space(foto_bytes):
    api_key = config('OCR_SPACE_API_KEY', default='')
    url_api = 'https://api.ocr.space/parse/image'
    
    # --- TRUCO: Comprimir la foto a menos de 1MB para el plan gratuito ---
    try:
        imagen = Image.open(io.BytesIO(foto_bytes))
        
        # Evitamos errores de compatibilidad si es un formato raro
        if imagen.mode in ("RGBA", "P"):
            imagen = imagen.convert("RGB")
            
        # Reducimos el tamaño y peso para que la API la acepte
        buffer = io.BytesIO()
        imagen.thumbnail((1200, 1200)) # Achicamos las dimensiones si es gigante
        imagen.save(buffer, format="JPEG", quality=50) 
        foto_optimizada = buffer.getvalue()
    except Exception as e:
        print(f"Alerta de compresión: {e}")
        foto_optimizada = foto_bytes # Si falla la compresión, mandamos la original
    # ---------------------------------------------------------------------

    payload = {
        'apikey': api_key,
        'language': 'eng',
        'isOverlayRequired': False,
        'scale': True,
        'OCREngine': '2'
    }
    
    try:
        # Enviamos la foto optimizada a OCR.space
        files = {'image': ('foto.jpg', foto_optimizada, 'image/jpeg')}
        respuesta = requests.post(url_api, files=files, data=payload, timeout=15)
        respuesta.raise_for_status()
        datos = respuesta.json()

        if datos.get('IsErroredOnProcessing'):
            return {'exitoso': False, 'error': datos.get('ErrorMessage', ['Error OCR'])[0]}

        resultados = datos.get('ParsedResults', [])
        if not resultados:
            return {'exitoso': False, 'error': 'No se detectó texto en la imagen.'}

        texto = resultados[0].get('ParsedText', '')
        
        print(f"--- TEXTO DETECTADO OCR.SPACE ---\n{texto}\n---------------------------------")
        return {'exitoso': True, 'texto': texto}
        
    except requests.exceptions.Timeout:
        return {'exitoso': False, 'error': 'El internet está lento. Intente de nuevo.'}
    except Exception as e:
        return {'exitoso': False, 'error': f'Error del servidor: {str(e)}'}
import re # Asegúrate de que siga arriba en tus imports

@login_required
@es_admin_tesorero_o_lector
def lectura_ocr_detectar(request):
    if request.method != 'POST':
        return JsonResponse({'exitoso': False, 'error': 'Método no permitido.'})
 
    foto = request.FILES.get('foto')
    if not foto:
        return JsonResponse({'exitoso': False, 'error': 'No se recibió ninguna foto.'})
 
    resultado_ocr = llamar_ocr_space(foto.read())
 
    if not resultado_ocr['exitoso']:
        return JsonResponse({'exitoso': False, 'error': resultado_ocr.get('error', 'No se procesó la imagen.')})
 
    texto_detectado = resultado_ocr.get('texto', '')
 
    # 1. Buscar medidor activo
    medidores_activos = Medidor.objects.select_related('socio').filter(estado='Activo')
    medidor_encontrado = None

    for m in medidores_activos:
        if m.numero_medidor and m.numero_medidor in texto_detectado:
            medidor_encontrado = m
            break
 
    if not medidor_encontrado:
        return JsonResponse({
            'exitoso': False,
            'error': 'No se detectó el número de serie del medidor. Intenta acercar la cámara.',
        })

    # =========================================================================
    # 2. HISTORIAL DEL SOCIO Y CÁLCULO DEL "ADIVINO"
    # =========================================================================
    lecturas_previas = Lectura.objects.filter(medidor=medidor_encontrado).order_by('-periodo')
    
    if lecturas_previas.exists():
        ultima_lectura = lecturas_previas.first()
        lectura_anterior_val = ultima_lectura.lectura_actual
        lectura_anterior_str = str(ultima_lectura.lectura_actual)
        periodo_sugerido = siguiente_periodo(ultima_lectura.periodo)
        ultimo_periodo_legible = periodo_nombre(ultima_lectura.periodo)
        
        # Calcular el consumo promedio de los últimos 3 meses de este socio
        ultimas_3 = lecturas_previas[:3]
        suma_consumo = sum([(l.lectura_actual - l.lectura_anterior) for l in ultimas_3])
        cantidad = ultimas_3.count()
        promedio_consumo = int(suma_consumo / cantidad) if cantidad > 0 else 6
        
        # Si la casa estuvo cerrada y el promedio da cero, le ponemos 6 por defecto
        if promedio_consumo < 2:
            promedio_consumo = 6
    else:
        # Si es un medidor totalmente nuevo
        lectura_anterior_val = Decimal('0.00')
        lectura_anterior_str = '0.00'
        periodo_sugerido = date.today().strftime('%Y-%m')
        ultimo_periodo_legible = 'Sin lecturas previas'
        promedio_consumo = 6 # Consumo base sugerido
 
    ya_existe = Lectura.objects.filter(medidor=medidor_encontrado, periodo=periodo_sugerido).exists()

    # =========================================================================
    # 3. EL CAZADOR Y AMPUTADOR INTELIGENTE
    # =========================================================================
    posible_lectura = ''
    numeros_encontrados = re.findall(r'\d+', texto_detectado)
    candidatos = []
    
    # Límite máximo de consumo permitido por IA antes de considerarlo "Basura"
    # (Si gasta más de 50 cubos en un mes, obligamos a que el lector lo escriba a mano)
    techo_maximo = lectura_anterior_val + 50 
    
    for num_str in numeros_encontrados:
        if num_str in medidor_encontrado.numero_medidor:
            continue
            
        # CASO A: La cámara solo leyó los negros (ej: "486" o "00486")
        try:
            val_raw = int(num_str)
            if lectura_anterior_val <= val_raw <= techo_maximo:
                candidatos.append(val_raw)
        except: pass
        
        # CASO B (EL AMPUTADOR): La cámara leyó negros y rojos juntos (ej: "0048612" o "48612")
        # Le quitamos los últimos 2 dígitos y vemos si el resto tiene sentido
        if len(num_str) >= 4:
            try:
                val_cortado = int(num_str[:-2]) 
                if lectura_anterior_val <= val_cortado <= techo_maximo:
                    candidatos.append(val_cortado)
            except: pass
            
    if candidatos:
        # Si encontró algo lógico, tomamos el que esté más cerca a la lectura anterior
        candidatos.sort()
        posible_lectura = str(candidatos[0])
    else:
        # EL ADIVINO: Si la foto estaba borrosa o el tambor a la mitad, sugerimos el promedio matemático
        sugerencia_matematica = int(lectura_anterior_val) + promedio_consumo
        posible_lectura = str(sugerencia_matematica)
 
    return JsonResponse({
        'exitoso': True,
        'numero_serie_detectado': medidor_encontrado.numero_medidor,
        'lectura_odometro_detectada': posible_lectura,
        'medidor': {
            'id': str(medidor_encontrado.pk),
            'numero_medidor': medidor_encontrado.numero_medidor,
            'socio_id': str(medidor_encontrado.socio.pk),
            'socio_nombre': medidor_encontrado.socio.nombre_completo,
            'socio_ci': medidor_encontrado.socio.ci,
            'manzano': medidor_encontrado.manzano or '',
            'parcela': medidor_encontrado.parcela or '',
        },
        'lectura_anterior': lectura_anterior_str,
        'periodo_sugerido': periodo_sugerido,
        'ultimo_periodo_legible': ultimo_periodo_legible,
        'ya_existe_periodo': ya_existe,
    })

@login_required
@es_admin_o_tesorero
def lectura_medidor_info(request, pk=None, medidor_id=None):
    from datetime import date
    
    # Resolver si viene como pk o como medidor_id desde urls.py
    id_buscar = pk or medidor_id
    
    medidor = get_object_or_404(Medidor.objects.select_related('socio'), pk=id_buscar)
    socio = medidor.socio

    # Detectar si el socio/medidor está inactivo o retirado
    socio_inactivo = socio.estado in ['INACTIVO', 'RETIRADO']
    medidor_inactivo = medidor.estado in ['Inactivo', 'Retirado']
    socio_bloqueado = socio_inactivo or medidor_inactivo

    # Obtener última lectura registrada
    ultima = medidor.lecturas.order_by('-fecha_lectura').first()
    
    # Calcular periodo sugerido
    hoy = date.today()
    periodo_sugerido = hoy.strftime('%Y-%m')
    ultimo_periodo_texto = "Sin lecturas previas"

    MESES = {
        1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL', 5: 'MAYO', 6: 'JUNIO',
        7: 'JULIO', 8: 'AGOSTO', 9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE'
    }

    if ultima and ultima.periodo:
        try:
            partes = ultima.periodo.split('-')
            anio, mes = int(partes[0]), int(partes[1])
            
            # Formato de texto ej: "JULIO 2026"
            if mes in MESES:
                ultimo_periodo_texto = f"{MESES[mes]} {anio}"
            else:
                ultimo_periodo_texto = ultima.periodo

            # Calcular mes siguiente sugerido
            mes += 1
            if mes > 12:
                mes = 1
                anio += 1
            periodo_sugerido = f'{anio:04d}-{mes:02d}'
        except Exception:
            ultimo_periodo_texto = ultima.periodo

    ya_existe = Lectura.objects.filter(medidor=medidor, periodo=periodo_sugerido).exists()

    return JsonResponse({
        'exitoso': True,
        'pk': str(medidor.pk),
        'numero_medidor': medidor.numero_medidor or 'Sin número',
        'socio_nombre': socio.nombre_completo,
        'socio_ci': socio.ci,
        'manzano': medidor.manzano or '—',
        'parcela': medidor.parcela or '—',
        'lectura_anterior': str(ultima.lectura_actual) if ultima else '0.00',
        'ultimo_periodo_legible': ultimo_periodo_texto,
        'periodo_sugerido': periodo_sugerido,
        'ya_existe_periodo': ya_existe,
        'socio_estado': socio.estado,
        'socio_bloqueado': socio_bloqueado,
    })
# =============================================================
# COBROS — SOLO ADMIN / TESORERO
# =============================================================

@login_required
@es_admin_o_tesorero
def cobros_lista(request):
    q = request.GET.get('q', '').strip()
    estado = request.GET.get('estado', '').strip()
    periodo = request.GET.get('periodo', '').strip()

    cobros = Cobro.objects.select_related(
        'socio',
        'lectura',
        'lectura__medidor'
    ).all()

    if q:
        cobros = cobros.filter(
            Q(numero_recibo__icontains=q) |
            Q(socio__nombre_completo__icontains=q) |
            Q(socio__ci__icontains=q) |
            Q(socio__codigo_cliente__icontains=q) |
            Q(lectura__medidor__numero_medidor__icontains=q) |
            Q(lectura__periodo__icontains=q)
        )

    if estado:
        cobros = cobros.filter(estado_pago=estado)

    if periodo:
        cobros = cobros.filter(lectura__periodo=periodo)

    estados = ['Pendiente', 'En Revision', 'Cancelado', 'Vencido']

    # Recalcula el recargo por atraso de cada cobro visible antes de
    # mostrarlo, para que la multa se refleje sola sin que nadie tenga
    # que aplicarla manualmente.
    for cobro in cobros:
        cobro.aplicar_recargo_automatico()

    return render(request, 'cobros/lista.html', {
        'cobros': cobros,
        'q': q,
        'estado': estado,
        'periodo': periodo,
        'estados': estados,
    })

@login_required
@es_admin_o_tesorero
def cobro_crear(request):
    """
    Pantalla de respaldo: el cobro normalmente se genera automáticamente
    al registrar la lectura. Esta vista solo cubre el caso excepcional
    de una lectura que haya quedado sin cobro asociado.
    """
    lecturas = Lectura.objects.select_related('medidor', 'medidor__socio').filter(recibo__isnull=True)

    if request.method == 'POST':
        lectura_id = request.POST.get('lectura')
        lectura = get_object_or_404(Lectura, pk=lectura_id)

        try:
            cobro = Cobro.objects.create(
                socio=lectura.medidor.socio,
                lectura=lectura,
                recargo_falta_pago=Decimal(request.POST.get('recargo_falta_pago') or '0'),
                otros=Decimal(request.POST.get('otros') or '0'),
            )
            marcados = {
                campo for campo in Cobro.CARGOS_ESPECIALES
                if request.POST.get(campo) == 'on'
            }
            cobro.aplicar_cargos_especiales(marcados)
            messages.success(request, f'Cobro N.º {cobro.numero_recibo} generado correctamente.')
            return redirect('cobros_lista')

        except Exception as e:
            messages.error(request, f'Error al generar cobro: {e}')

    return render(request, 'cobros/generar.html', {
        'accion': 'Crear',
        'lecturas': lecturas,
    })


@login_required
@es_admin_o_tesorero
def cobro_detalle(request, pk):
    cobro = get_object_or_404(
        Cobro.objects.select_related('socio', 'lectura', 'lectura__medidor'),
        pk=pk
    )
    cobro.aplicar_recargo_automatico()

    pagos = cobro.pagos.all()
    total_pagado = pagos.aggregate(total=Sum('monto_pagado'))['total'] or Decimal('0.00')
    saldo = cobro.monto_total - total_pagado

    return render(request, 'cobros/detalle.html', {
        'cobro': cobro,
        'pagos': pagos,
        'total_pagado': total_pagado,
        'saldo': saldo,
    })

@login_required
@es_admin_o_tesorero
def cobro_editar_cargos(request, pk):
    """
    Permite activar/desactivar los cargos especiales de un cobro ya
    generado (reconexión, limpieza de tanque, instalación clandestina,
    multa por alteración, falta a asamblea) mediante checkboxes. El
    monto de cada uno se toma siempre de la tarifa activa, nunca se
    escribe a mano.
    """
    cobro = get_object_or_404(
        Cobro.objects.select_related('socio', 'lectura', 'lectura__medidor'),
        pk=pk
    )

    if request.method == 'POST':
        marcados = {
            campo for campo in Cobro.CARGOS_ESPECIALES
            if request.POST.get(campo) == 'on'
        }
        cobro.aplicar_cargos_especiales(marcados)
        messages.success(request, 'Cargos del cobro actualizados correctamente.')
        return redirect('cobro_detalle', pk=cobro.pk)

    tarifa = Tarifa.objects.filter(activa=True).order_by('-id_tarifa').first()

    return render(request, 'cobros/editar_cargos.html', {
        'cobro': cobro,
        'tarifa': tarifa,
    })

@login_required
@es_admin_o_tesorero
def cobro_imprimir(request, pk):
    cobro = get_object_or_404(
        Cobro.objects.select_related(
            'socio',
            'lectura',
            'lectura__medidor'
        ),
        pk=pk
    )
    cobro.aplicar_recargo_automatico()

    pagos = cobro.pagos.all()
    total_pagado = pagos.aggregate(total=Sum('monto_pagado'))['total'] or Decimal('0.00')
    saldo = cobro.monto_total - total_pagado

    return render(request, 'cobros/imprimir.html', {
        'cobro': cobro,
        'pagos': pagos,
        'total_pagado': total_pagado,
        'saldo': saldo,
    })
# =============================================================
# PAGOS — SOLO ADMIN / TESORERO
# =============================================================
@login_required
@es_admin_o_tesorero
def pagos_lista(request):
    q = request.GET.get('q', '').strip()

    pagos = Pago.objects.select_related('recibo', 'recibo__socio').all()

    if q:
        pagos = pagos.filter(
            Q(recibo__numero_recibo__icontains=q) |
            Q(recibo__socio__nombre_completo__icontains=q) |
            Q(recibo__socio__ci__icontains=q) |
            Q(metodo_pago__icontains=q)
        )

    return render(request, 'pagos/lista.html', {
        'pagos': pagos,
        'q': q,
    })


@login_required
@es_admin_o_tesorero
def pago_registrar(request, cobro_pk):
    cobro = get_object_or_404(Cobro, pk=cobro_pk)
    total_pagado = cobro.pagos.aggregate(t=Sum('monto_pagado'))['t'] or Decimal('0.00')
    saldo = cobro.monto_total - total_pagado

    if request.method == 'POST':
        from decimal import InvalidOperation
        try:
            monto = Decimal(request.POST.get('monto_pagado', '0'))
            metodo = request.POST.get('metodo_pago', 'efectivo')
            foto = request.FILES.get('foto_comprobante')

            if monto <= 0:
                messages.error(request, 'El monto debe ser mayor a cero.')
            elif monto > saldo:
                messages.error(request, f'El monto no puede superar el saldo de Bs {saldo}.')
            else:
                Pago.objects.create(
                    recibo=cobro,
                    monto_pagado=monto,
                    metodo_pago=metodo,
                    foto_comprobante=foto,
                    registrado_por=request.user,
                )
                messages.success(request, f'Pago de Bs {monto} registrado correctamente.')
                return redirect('cobro_detalle', pk=cobro.pk)
        except InvalidOperation:
            messages.error(request, 'Monto inválido.')

    return render(request, 'pagos/form.html', {
        'cobro': cobro,
        'saldo': saldo,
    })

@login_required
@es_admin_o_tesorero
def pago_detalle(request, pk):
    pago = get_object_or_404(
        Pago.objects.select_related('recibo', 'recibo__socio'),
        pk=pk
    )

    return render(request, 'pagos/detalle.html', {'pago': pago})
# =============================================================
# AGREGA ESTAS FUNCIONES EN views_web.py
# =============================================================

def monto_a_literal(monto):
    """
    Convierte un monto decimal a su representación en letras.
    Ejemplo: 60.50 → "SESENTA CON 50/100 BOLIVIANOS"
    Cubre hasta 999.99 Bs (suficiente para recibos de agua).
    """
    from decimal import Decimal

    unidades = ['', 'UNO', 'DOS', 'TRES', 'CUATRO', 'CINCO',
                'SEIS', 'SIETE', 'OCHO', 'NUEVE', 'DIEZ',
                'ONCE', 'DOCE', 'TRECE', 'CATORCE', 'QUINCE',
                'DIECISÉIS', 'DIECISIETE', 'DIECIOCHO', 'DIECINUEVE', 'VEINTE',
                'VEINTIUNO', 'VEINTIDÓS', 'VEINTITRÉS', 'VEINTICUATRO', 'VEINTICINCO',
                'VEINTISÉIS', 'VEINTISIETE', 'VEINTIOCHO', 'VEINTINUEVE']

    decenas = ['', '', 'VEINTE', 'TREINTA', 'CUARENTA', 'CINCUENTA',
               'SESENTA', 'SETENTA', 'OCHENTA', 'NOVENTA']

    centenas = ['', 'CIEN', 'DOSCIENTOS', 'TRESCIENTOS', 'CUATROCIENTOS',
                'QUINIENTOS', 'SEISCIENTOS', 'SETECIENTOS', 'OCHOCIENTOS', 'NOVECIENTOS']

    try:
        monto = Decimal(str(monto))
        entero = int(monto)
        centavos = int(round((monto - entero) * 100))
    except Exception:
        return 'MONTO INVÁLIDO'

    def decenas_letras(n):
        if n < 29:
            return unidades[n]
        d, u = divmod(n, 10)
        if u == 0:
            return decenas[d]
        return f"{decenas[d]} Y {unidades[u]}"

    def numero_letras(n):
        if n == 0:
            return 'CERO'
        if n < 100:
            return decenas_letras(n)
        c, resto = divmod(n, 100)
        if resto == 0:
            return centenas[c]
        if c == 1:
            return f"CIENTO {decenas_letras(resto)}"
        return f"{centenas[c]} {decenas_letras(resto)}"

    letras = numero_letras(entero)
    return f"{letras} CON {centavos:02d}/100 BOLIVIANOS"


# =============================================================
# VISTA PARA IMPRIMIR EL TICKET TÉRMICO
# Reemplaza o agrega junto a tu vista cobro_imprimir existente
# =============================================================

@login_required
@es_admin_o_tesorero
def cobro_imprimir_termico(request, pk):
    cobro = get_object_or_404(
        Cobro.objects.select_related('socio', 'lectura', 'lectura__medidor'),
        pk=pk
    )
    cobro.aplicar_recargo_automatico()

    pagos = cobro.pagos.all()
    total_pagado = pagos.aggregate(total=Sum('monto_pagado'))['total'] or Decimal('0.00')
    saldo = cobro.monto_total - total_pagado

    return render(request, 'cobros/imprimir_termico.html', {
        'cobro': cobro,
        'pagos': pagos,
        'total_pagado': total_pagado,
        'saldo': saldo,
        'monto_literal': monto_a_literal(cobro.monto_total),
        'ahora': timezone.now(),
    })

# =============================================================
# TARIFAS — SOLO ADMIN / TESORERO
# =============================================================

@login_required
@es_solo_admin
def tarifas_lista(request):
    tarifas = Tarifa.objects.all().order_by('-activa', '-id_tarifa')
    return render(request, 'tarifas/lista.html', {'tarifas': tarifas})

@login_required
@es_solo_admin
def tarifa_crear(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        costo_por_cubo = Decimal(request.POST.get('costo_por_cubo') or '0')
        cuota_fija = Decimal(request.POST.get('cuota_fija') or '0')
        multa_atraso = Decimal(request.POST.get('multa_atraso') or '0')
        dias_gracia = int(request.POST.get('dias_gracia') or 0)
        costo_reconexion = Decimal(request.POST.get('costo_reconexion') or '0')
        costo_limpieza_tanque = Decimal(request.POST.get('costo_limpieza_tanque') or '0')
        costo_instalacion_clandestina = Decimal(request.POST.get('costo_instalacion_clandestina') or '0')
        costo_multa_alteracion = Decimal(request.POST.get('costo_multa_alteracion') or '0')
        costo_falta_asamblea = Decimal(request.POST.get('costo_falta_asamblea') or '0')
        costo_afiliacion = Decimal(request.POST.get('costo_afiliacion') or '0')
        activa = request.POST.get('activa') == 'on'

        montos = [
            costo_por_cubo, cuota_fija, multa_atraso,
            costo_reconexion, costo_limpieza_tanque,
            costo_instalacion_clandestina, costo_multa_alteracion,
            costo_falta_asamblea, costo_afiliacion,
        ]

        if not nombre:
            messages.error(request, 'El nombre de la tarifa es obligatorio.')
        elif any(m < 0 for m in montos):
            messages.error(request, 'Los montos no pueden ser negativos.')
        elif dias_gracia < 0:
            messages.error(request, 'Los días de gracia no pueden ser negativos.')
        else:
            if activa:
                Tarifa.objects.update(activa=False)

            Tarifa.objects.create(
                nombre=nombre,
                costo_por_cubo=costo_por_cubo,
                cuota_fija=cuota_fija,
                multa_atraso=multa_atraso,
                dias_gracia=dias_gracia,
                costo_reconexion=costo_reconexion,
                costo_limpieza_tanque=costo_limpieza_tanque,
                costo_instalacion_clandestina=costo_instalacion_clandestina,
                costo_multa_alteracion=costo_multa_alteracion,
                costo_falta_asamblea=costo_falta_asamblea,
                costo_afiliacion=costo_afiliacion,
                activa=activa,
            )

            messages.success(request, 'Tarifa creada correctamente.')
            return redirect('tarifas_lista')

    return render(request, 'tarifas/form.html', {
        'accion': 'Crear'
    })

@login_required
@es_solo_admin
def tarifa_editar(request, pk):
    tarifa = get_object_or_404(Tarifa, pk=pk)

    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        costo_por_cubo = Decimal(request.POST.get('costo_por_cubo') or '0')
        cuota_fija = Decimal(request.POST.get('cuota_fija') or '0')
        multa_atraso = Decimal(request.POST.get('multa_atraso') or '0')
        dias_gracia = int(request.POST.get('dias_gracia') or 0)
        costo_reconexion = Decimal(request.POST.get('costo_reconexion') or '0')
        costo_limpieza_tanque = Decimal(request.POST.get('costo_limpieza_tanque') or '0')
        costo_instalacion_clandestina = Decimal(request.POST.get('costo_instalacion_clandestina') or '0')
        costo_multa_alteracion = Decimal(request.POST.get('costo_multa_alteracion') or '0')
        costo_falta_asamblea = Decimal(request.POST.get('costo_falta_asamblea') or '0')
        costo_afiliacion = Decimal(request.POST.get('costo_afiliacion') or '0')
        activa = request.POST.get('activa') == 'on'

        montos = [
            costo_por_cubo, cuota_fija, multa_atraso,
            costo_reconexion, costo_limpieza_tanque,
            costo_instalacion_clandestina, costo_multa_alteracion,
            costo_falta_asamblea, costo_afiliacion,
        ]

        if not nombre:
            messages.error(request, 'El nombre de la tarifa es obligatorio.')
        elif any(m < 0 for m in montos):
            messages.error(request, 'Los montos no pueden ser negativos.')
        elif dias_gracia < 0:
            messages.error(request, 'Los días de gracia no pueden ser negativos.')
        else:
            if activa:
                Tarifa.objects.exclude(pk=tarifa.pk).update(activa=False)

            tarifa.nombre = nombre
            tarifa.costo_por_cubo = costo_por_cubo
            tarifa.cuota_fija = cuota_fija
            tarifa.multa_atraso = multa_atraso
            tarifa.dias_gracia = dias_gracia
            tarifa.costo_reconexion = costo_reconexion
            tarifa.costo_limpieza_tanque = costo_limpieza_tanque
            tarifa.costo_instalacion_clandestina = costo_instalacion_clandestina
            tarifa.costo_multa_alteracion = costo_multa_alteracion
            tarifa.costo_falta_asamblea = costo_falta_asamblea
            tarifa.costo_afiliacion = costo_afiliacion
            tarifa.activa = activa
            tarifa.save()

            messages.success(request, 'Tarifa actualizada correctamente.')
            return redirect('tarifas_lista')

    return render(request, 'tarifas/form.html', {
        'accion': 'Editar',
        'tarifa': tarifa,
    })

# =============================================================
# USUARIOS — SOLO ADMIN / TESORERO
# =============================================================

@login_required
@es_solo_admin
def usuarios_lista(request):
    q = request.GET.get('q', '').strip()
    usuarios = Usuario.objects.all().order_by('username')

    if q:
        usuarios = usuarios.filter(
            Q(username__icontains=q) |
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(rol__icontains=q)
        )

    return render(request, 'usuarios/lista.html', {
        'usuarios': usuarios,
        'q': q,
    })

@login_required
@es_solo_admin
def usuario_crear(request):
    socios = Socio.objects.all().order_by('nombre_completo')
    roles = Usuario.ROL_CHOICES

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        ci = request.POST.get('ci', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        rol = request.POST.get('rol', '').strip()
        socio_id = request.POST.get('socio') or None
        activo = request.POST.get('activo') == 'on'

        if not username:
            messages.error(request, 'El nombre de usuario es obligatorio.')
        elif not password:
            messages.error(request, 'La contraseña es obligatoria.')
        elif not rol:
            messages.error(request, 'Debe seleccionar un rol.')
        elif Usuario.objects.filter(username=username).exists():
            messages.error(request, 'Ya existe un usuario con ese nombre.')
        elif ci and Usuario.objects.filter(ci=ci).exists():
            messages.error(request, 'Ya existe un usuario con ese CI.')
        else:
            usuario = Usuario.objects.create_user(
                username=username,
                password=password,
                first_name=first_name,
                last_name=last_name,
                email=email,
            )

            usuario.rol = rol
            usuario.ci = ci or None
            usuario.telefono = telefono or None
            usuario.activo = activo
            usuario.is_active = activo

            if rol == 'socio' and socio_id:
                socio = get_object_or_404(Socio, pk=socio_id)
                socio.usuario = usuario
                socio.save()

            usuario.save()

            messages.success(request, 'Usuario creado correctamente.')
            return redirect('usuarios_lista')

    return render(request, 'usuarios/form.html', {
        'accion': 'Crear',
        'socios': socios,
        'roles': roles,
    })
@login_required
@es_admin_o_tesorero
def usuario_editar(request, pk):
    usuario = get_object_or_404(Usuario, pk=pk)
    socios = Socio.objects.all().order_by('nombre_completo')
    roles = Usuario.ROL_CHOICES

    socio_asociado = Socio.objects.filter(usuario=usuario).first()

    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        ci = request.POST.get('ci', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        rol = request.POST.get('rol', '').strip()
        socio_id = request.POST.get('socio') or None
        activo = request.POST.get('activo') == 'on'

        if not rol:
            messages.error(request, 'Debe seleccionar un rol.')
        elif ci and Usuario.objects.filter(ci=ci).exclude(pk=usuario.pk).exists():
            messages.error(request, 'Ya existe otro usuario con ese CI.')
        else:
            usuario.first_name = first_name
            usuario.last_name = last_name
            usuario.email = email
            usuario.ci = ci or None
            usuario.telefono = telefono or None
            usuario.rol = rol
            usuario.activo = activo
            usuario.is_active = activo
            usuario.save()

            # Quitar asociación anterior si existía
            Socio.objects.filter(usuario=usuario).update(usuario=None)

            # Asociar socio si el rol es socio
            if rol == 'socio' and socio_id:
                socio = get_object_or_404(Socio, pk=socio_id)
                socio.usuario = usuario
                socio.save()

            messages.success(request, 'Usuario actualizado correctamente.')
            return redirect('usuarios_lista')

    return render(request, 'usuarios/form.html', {
        'accion': 'Editar',
        'usuario_obj': usuario,
        'socios': socios,
        'roles': roles,
        'socio_asociado': socio_asociado,
    })

@login_required
@es_solo_admin
def usuario_reset_password(request, pk):
    usuario = get_object_or_404(Usuario, pk=pk)

    if request.method == 'POST':
        nueva = request.POST.get('password_nuevo', '').strip()
        confirmar = request.POST.get('password_confirmar', '').strip()

        if not nueva:
            messages.error(request, 'La nueva contraseña es obligatoria.')
        elif nueva != confirmar:
            messages.error(request, 'Las contraseñas no coinciden.')
        elif len(nueva) < 6:
            messages.error(request, 'La contraseña debe tener al menos 6 caracteres.')
        else:
            usuario.set_password(nueva)
            usuario.save()
            messages.success(request, 'Contraseña restablecida correctamente.')
            return redirect('usuarios_lista')

    return render(request, 'usuarios/reset_password.html', {
        'usuario_obj': usuario,
    })
# =============================================================
# REPORTES — SOLO ADMIN / TESORERO
# =============================================================

MESES_NOMBRE = {
    '01': 'ENERO',
    '02': 'FEBRERO',
    '03': 'MARZO',
    '04': 'ABRIL',
    '05': 'MAYO',
    '06': 'JUNIO',
    '07': 'JULIO',
    '08': 'AGOSTO',
    '09': 'SEPTIEMBRE',
    '10': 'OCTUBRE',
    '11': 'NOVIEMBRE',
    '12': 'DICIEMBRE',
}


def nombre_periodo(periodo):
    try:
        anio, mes = periodo.split('-')
        return f"{MESES_NOMBRE.get(mes, mes)} {anio}"
    except Exception:
        return periodo


def _metadata_reporte(request):
    """Datos comunes de auditoría para el encabezado de cada reporte."""
    return {
        'generado_por': request.user.get_full_name(),
        'fecha_generacion': timezone.now(),
    }


@login_required
@es_admin_o_tesorero
def reportes_view(request):
    """
    Menú principal de reportes.
    No es dashboard; solo muestra accesos a reportes imprimibles.
    """
    return render(request, 'reportes/index.html')


@login_required
@es_admin_o_tesorero
def reporte_deudas(request):
    estado = request.GET.get('estado', '').strip()
    periodo = request.GET.get('periodo', '').strip()
    q = request.GET.get('q', '').strip()

    estados_deuda = ['Pendiente', 'En Revision', 'En Revisión', 'Vencido']

    recibos = Cobro.objects.select_related(
        'socio',
        'lectura',
        'lectura__medidor'
    ).filter(
        estado_pago__in=estados_deuda
    )

    if estado:
        recibos = recibos.filter(estado_pago=estado)

    if periodo:
        recibos = recibos.filter(lectura__periodo=periodo)

    if q:
        recibos = recibos.filter(
            Q(socio__nombre_completo__icontains=q) |
            Q(socio__ci__icontains=q) |
            Q(socio__codigo_cliente__icontains=q) |
            Q(lectura__medidor__numero_medidor__icontains=q) |
            Q(numero_recibo__icontains=q)
        )

    recibos = recibos.order_by('socio__nombre_completo', '-fecha_emision')

    total_deuda = recibos.aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')
    cantidad = recibos.count()

    return render(request, 'reportes/deudas.html', {
        'recibos': recibos,
        'total_deuda': total_deuda,
        'cantidad': cantidad,
        'estado': estado,
        'periodo': periodo,
        'periodo_nombre': nombre_periodo(periodo) if periodo else '',
        'q': q,
        'estados': estados_deuda,
        **_metadata_reporte(request),
    })


@login_required
@es_admin_o_tesorero
def reporte_recaudacion(request):
    fecha_inicio = request.GET.get('fecha_inicio', '').strip()
    fecha_fin = request.GET.get('fecha_fin', '').strip()
    metodo = request.GET.get('metodo', '').strip()
    q = request.GET.get('q', '').strip()

    pagos = Pago.objects.select_related(
        'recibo',
        'recibo__socio',
        'recibo__lectura',
        'recibo__lectura__medidor'
    ).all()

    if fecha_inicio:
        fi = parse_date(fecha_inicio)
        if fi:
            pagos = pagos.filter(fecha_pago__date__gte=fi)

    if fecha_fin:
        ff = parse_date(fecha_fin)
        if ff:
            pagos = pagos.filter(fecha_pago__date__lte=ff)

    if metodo:
        pagos = pagos.filter(metodo_pago=metodo)

    if q:
        pagos = pagos.filter(
            Q(recibo__numero_recibo__icontains=q) |
            Q(recibo__socio__nombre_completo__icontains=q) |
            Q(recibo__socio__ci__icontains=q) |
            Q(recibo__lectura__medidor__numero_medidor__icontains=q)
        )

    pagos = pagos.order_by('-fecha_pago')

    total_recaudado = pagos.aggregate(total=Sum('monto_pagado'))['total'] or Decimal('0.00')
    cantidad = pagos.count()

    # Afiliaciones (acción de agua) pagadas en el mismo rango de fechas.
    # No pasan por Pago (no tienen método de pago), por eso se calculan aparte
    # y se suman al total general de lo recaudado.
    afiliaciones = Afiliacion.objects.select_related('socio', 'medidor').all()

    if fecha_inicio:
        fi = parse_date(fecha_inicio)
        if fi:
            afiliaciones = afiliaciones.filter(fecha_pago__gte=fi)

    if fecha_fin:
        ff = parse_date(fecha_fin)
        if ff:
            afiliaciones = afiliaciones.filter(fecha_pago__lte=ff)

    if q:
        afiliaciones = afiliaciones.filter(
            Q(socio__nombre_completo__icontains=q) |
            Q(socio__ci__icontains=q) |
            Q(medidor__numero_medidor__icontains=q)
        )

    total_afiliaciones = afiliaciones.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    cantidad_afiliaciones = afiliaciones.count()
    total_general = total_recaudado + total_afiliaciones

    return render(request, 'reportes/recaudacion.html', {
        'pagos': pagos,
        'total_recaudado': total_recaudado,
        'cantidad': cantidad,
        'total_afiliaciones': total_afiliaciones,
        'cantidad_afiliaciones': cantidad_afiliaciones,
        'total_general': total_general,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'metodo': metodo,
        'q': q,
        'metodos': Pago.METODO_CHOICES,
        **_metadata_reporte(request),
    })


@login_required
@es_admin_o_tesorero
def reporte_mensual(request):
    periodo = request.GET.get('periodo', '').strip()

    if not periodo:
        hoy = date.today()
        periodo = f'{hoy.year}-{hoy.month:02d}'

    recibos = Cobro.objects.select_related(
        'socio',
        'lectura',
        'lectura__medidor'
    ).filter(
        lectura__periodo=periodo
    ).order_by('socio__nombre_completo')

    pagos = Pago.objects.select_related(
        'recibo',
        'recibo__socio'
    ).filter(
        recibo__lectura__periodo=periodo
    )

    total_emitido = recibos.aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')
    total_pagado = pagos.aggregate(total=Sum('monto_pagado'))['total'] or Decimal('0.00')
    total_pendiente = total_emitido - total_pagado

    consumo_total = Decimal('0.00')
    for recibo in recibos:
        consumo_total += recibo.lectura.consumo_cubos

    # Afiliaciones pagadas dentro del mismo mes/año del periodo seleccionado.
    anio_periodo, mes_periodo = periodo.split('-')
    afiliaciones = Afiliacion.objects.filter(
        fecha_pago__year=int(anio_periodo),
        fecha_pago__month=int(mes_periodo)
    )
    total_afiliaciones = afiliaciones.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    total_pagado_general = total_pagado + total_afiliaciones

    return render(request, 'reportes/mensual.html', {
        'periodo': periodo,
        'periodo_nombre': nombre_periodo(periodo),
        'recibos': recibos,
        'total_emitido': total_emitido,
        'total_pagado': total_pagado,
        'total_pendiente': total_pendiente,
        'consumo_total': consumo_total,
        'cantidad_recibos': recibos.count(),
        'afiliaciones': afiliaciones,
        'total_afiliaciones': total_afiliaciones,
        'cantidad_afiliaciones': afiliaciones.count(),
        'total_pagado_general': total_pagado_general,
        **_metadata_reporte(request),
    })


@login_required
@es_admin_o_tesorero
def reporte_anual(request):
    anio = request.GET.get('anio', '').strip()

    if not anio:
        anio = str(date.today().year)

    recibos = Cobro.objects.select_related(
        'lectura'
    ).filter(
        lectura__periodo__startswith=anio
    )

    pagos = Pago.objects.select_related(
        'recibo',
        'recibo__lectura'
    ).filter(
        recibo__lectura__periodo__startswith=anio
    )

    resumen_meses = []
    total_afiliaciones_anual = Decimal('0.00')

    for numero_mes in range(1, 13):
        mes = f'{numero_mes:02d}'
        periodo = f'{anio}-{mes}'

        recibos_mes = recibos.filter(lectura__periodo=periodo)
        pagos_mes = pagos.filter(recibo__lectura__periodo=periodo)
        afiliaciones_mes = Afiliacion.objects.filter(
            fecha_pago__year=int(anio),
            fecha_pago__month=numero_mes
        )

        total_emitido = recibos_mes.aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')
        total_pagado = pagos_mes.aggregate(total=Sum('monto_pagado'))['total'] or Decimal('0.00')
        total_afiliaciones_mes = afiliaciones_mes.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
        total_afiliaciones_anual += total_afiliaciones_mes

        consumo_mes = Decimal('0.00')
        for recibo in recibos_mes:
            consumo_mes += recibo.lectura.consumo_cubos

        resumen_meses.append({
            'periodo': periodo,
            'mes_nombre': MESES_NOMBRE[mes],
            'recibos': recibos_mes.count(),
            'emitido': total_emitido,
            'pagado': total_pagado,
            'pendiente': total_emitido - total_pagado,
            'consumo': consumo_mes,
            'afiliaciones': total_afiliaciones_mes,
        })

    total_emitido_anual = recibos.aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')
    total_pagado_anual = pagos.aggregate(total=Sum('monto_pagado'))['total'] or Decimal('0.00')

    # Detalle de quién pagó cada afiliación durante la gestión (para mostrar
    # nombre y no solo el total agregado por mes).
    afiliaciones_anual = Afiliacion.objects.select_related(
        'socio', 'medidor'
    ).filter(
        fecha_pago__year=int(anio)
    ).order_by('-fecha_pago')

    return render(request, 'reportes/anual.html', {
        'anio': anio,
        'resumen_meses': resumen_meses,
        'total_emitido_anual': total_emitido_anual,
        'total_pagado_anual': total_pagado_anual,
        'total_pendiente_anual': total_emitido_anual - total_pagado_anual,
        'total_afiliaciones_anual': total_afiliaciones_anual,
        'total_general_anual': total_pagado_anual + total_afiliaciones_anual,
        'afiliaciones_anual': afiliaciones_anual,
        **_metadata_reporte(request),
    })


@login_required
@es_admin_o_tesorero
def reporte_multas(request):
    periodo = request.GET.get('periodo', '').strip()
    anio = request.GET.get('anio', '').strip()

    recibos = Cobro.objects.select_related(
        'socio',
        'lectura',
        'lectura__medidor'
    ).all()

    if periodo:
        recibos = recibos.filter(lectura__periodo=periodo)

    if anio:
        recibos = recibos.filter(lectura__periodo__startswith=anio)

    recibos = recibos.filter(
        Q(recargo_falta_pago__gt=0) |
        Q(instalacion_clandestina__gt=0) |
        Q(multa_alteracion__gt=0) |
        Q(reconexion__gt=0) |
        Q(limpieza_tanque__gt=0) |
        Q(falta_asamblea__gt=0) |
        Q(otros__gt=0)
    ).order_by('-fecha_emision')

    totales = {
        'recargo_falta_pago': recibos.aggregate(t=Sum('recargo_falta_pago'))['t'] or Decimal('0.00'),
        'instalacion_clandestina': recibos.aggregate(t=Sum('instalacion_clandestina'))['t'] or Decimal('0.00'),
        'multa_alteracion': recibos.aggregate(t=Sum('multa_alteracion'))['t'] or Decimal('0.00'),
        'reconexion': recibos.aggregate(t=Sum('reconexion'))['t'] or Decimal('0.00'),
        'limpieza_tanque': recibos.aggregate(t=Sum('limpieza_tanque'))['t'] or Decimal('0.00'),
        'falta_asamblea': recibos.aggregate(t=Sum('falta_asamblea'))['t'] or Decimal('0.00'),
        'otros': recibos.aggregate(t=Sum('otros'))['t'] or Decimal('0.00'),
    }

    total_general = sum(totales.values(), Decimal('0.00'))

    return render(request, 'reportes/multas.html', {
        'recibos': recibos,
        'periodo': periodo,
        'periodo_nombre': nombre_periodo(periodo) if periodo else '',
        'anio': anio,
        'totales': totales,
        'total_general': total_general,
        **_metadata_reporte(request),
    })


@login_required
@es_admin_o_tesorero
def reporte_afiliaciones(request):
    tipo = request.GET.get('tipo', '').strip()
    fecha_inicio = request.GET.get('fecha_inicio', '').strip()
    fecha_fin = request.GET.get('fecha_fin', '').strip()
    q = request.GET.get('q', '').strip()

    afiliaciones = Afiliacion.objects.select_related(
        'socio', 'medidor', 'registrado_por'
    ).all()

    if tipo:
        afiliaciones = afiliaciones.filter(tipo=tipo)

    if fecha_inicio:
        fi = parse_date(fecha_inicio)
        if fi:
            afiliaciones = afiliaciones.filter(fecha_pago__gte=fi)

    if fecha_fin:
        ff = parse_date(fecha_fin)
        if ff:
            afiliaciones = afiliaciones.filter(fecha_pago__lte=ff)

    if q:
        afiliaciones = afiliaciones.filter(
            Q(socio__nombre_completo__icontains=q) |
            Q(socio__ci__icontains=q) |
            Q(medidor__numero_medidor__icontains=q)
        )

    afiliaciones = afiliaciones.order_by('-fecha_pago')

    total_afiliaciones = afiliaciones.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    cantidad = afiliaciones.count()
    cantidad_nuevo = afiliaciones.filter(tipo='NUEVO').count()
    cantidad_transferencia = afiliaciones.filter(tipo='TRANSFERENCIA').count()

    return render(request, 'reportes/afiliaciones.html', {
        'afiliaciones': afiliaciones,
        'total_afiliaciones': total_afiliaciones,
        'cantidad': cantidad,
        'cantidad_nuevo': cantidad_nuevo,
        'cantidad_transferencia': cantidad_transferencia,
        'tipo': tipo,
        'tipos': Afiliacion.TIPO_CHOICES,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'q': q,
        **_metadata_reporte(request),
    })

# =============================================================
# QRS GENÉRICOS (CRUD WEB) — SOLO ADMIN / TESORERO
# =============================================================

@login_required
@es_admin_o_tesorero
def qrs_lista(request):
    qrs = QRGenerico.objects.all().order_by('monto')
    return render(request, 'qrs/lista.html', {'qrs': qrs})

@login_required
@es_admin_o_tesorero
def qr_crear(request):
    if request.method == 'POST':
        monto = request.POST.get('monto', '').strip()
        imagen = request.FILES.get('imagen_qr')
        activo = request.POST.get('activo') == 'on'
        sin_monto = request.POST.get('sin_monto') == 'on'

        if not imagen:
            messages.error(request, 'La imagen es obligatoria.')
        elif not sin_monto and not monto:
            messages.error(request, 'El monto es obligatorio, o marca "QR genérico sin monto".')
        elif not sin_monto and QRGenerico.objects.filter(monto=monto).exists():
            messages.error(request, f'Ya existe un QR registrado para el monto de {monto} Bs.')
        elif sin_monto and activo and QRGenerico.objects.filter(monto__isnull=True, activo=True).exists():
            messages.error(request, 'Ya existe un QR genérico (sin monto) activo. Desactívalo antes de crear otro.')
        else:
            try:
                QRGenerico.objects.create(
                    monto=None if sin_monto else monto,
                    imagen_qr=imagen,
                    activo=activo
                )
                if sin_monto:
                    messages.success(request, 'QR genérico (sin monto) creado exitosamente.')
                else:
                    messages.success(request, f'QR de {monto} Bs creado exitosamente.')
                return redirect('qrs_lista')
            except Exception as e:
                messages.error(request, f'Error al guardar: {e}')

    return render(request, 'qrs/form.html', {'accion': 'Crear'})

@login_required
@es_admin_o_tesorero
def qr_editar(request, pk):
    qr = get_object_or_404(QRGenerico, pk=pk)

    if request.method == 'POST':
        monto = request.POST.get('monto', '').strip()
        imagen = request.FILES.get('imagen_qr')
        activo = request.POST.get('activo') == 'on'
        sin_monto = request.POST.get('sin_monto') == 'on'

        if not sin_monto and not monto:
            messages.error(request, 'El monto es obligatorio, o marca "QR genérico sin monto".')
        elif not sin_monto and QRGenerico.objects.filter(monto=monto).exclude(pk=qr.pk).exists():
            messages.error(request, f'Ya existe otro QR con el monto de {monto} Bs.')
        elif sin_monto and activo and QRGenerico.objects.filter(monto__isnull=True, activo=True).exclude(pk=qr.pk).exists():
            messages.error(request, 'Ya existe otro QR genérico (sin monto) activo. Desactívalo antes de activar este.')
        else:
            try:
                qr.monto = None if sin_monto else monto
                qr.activo = activo
                if imagen:  # Solo actualiza la imagen si se subió una nueva
                    qr.imagen_qr = imagen
                qr.save()
                messages.success(request, 'QR actualizado correctamente.')
                return redirect('qrs_lista')
            except Exception as e:
                messages.error(request, f'Error al actualizar: {e}')

    return render(request, 'qrs/form.html', {
        'accion': 'Editar',
        'qr': qr
    })

@login_required
@es_admin_o_tesorero
def qr_eliminar(request, pk):
    qr = get_object_or_404(QRGenerico, pk=pk)
    if request.method == 'POST':
        qr.delete()
        messages.success(request, 'QR eliminado correctamente.')
        return redirect('qrs_lista')
    return render(request, 'qrs/confirmar_eliminar.html', {'qr': qr})
# =============================================================
# ALIAS PARA COMPATIBILIDAD CON URLS ANTIGUAS
# =============================================================

cobro_generar = cobro_crear

# =============================================================
# AGREGA ESTAS FUNCIONES EN views_web.py
# Backup completo del sistema en Excel — descargable desde el panel
# =============================================================
@login_required
@es_admin_o_tesorero
def backup_vista(request):
    """Página principal de backup e importación de datos."""
    stats = {
        'socios': Socio.objects.count(),
        'medidores': Medidor.objects.count(),
        'lecturas': Lectura.objects.count(),
        'cobros': Cobro.objects.count(),
        'pagos': Pago.objects.count(),
        'afiliaciones': Afiliacion.objects.count(),
        'total_recaudado': Pago.objects.aggregate(t=Sum('monto_pagado'))['t'] or Decimal('0.00'),
        'ahora': timezone.now(),
    }
    return render(request, 'backup/index.html', stats)

@login_required
@es_admin_o_tesorero
def backup_excel(request):
    """
    Genera un archivo Excel con TODOS los datos del sistema.
    Incluye hoja de Resumen Ejecutivo con estadísticas globales.
    Requiere confirmación por contraseña.
    """
    if request.method != 'POST':
        messages.error(request, 'Método no permitido. Debe confirmar con su contraseña.')
        return redirect('backup_vista')

    password_confirmacion = request.POST.get('password_confirmacion', '')
    if not request.user.check_password(password_confirmacion):
        messages.error(request, 'Contraseña incorrecta. No se pudo generar el backup.')
        return redirect('backup_vista')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Quitar hoja por defecto

    # Estilos
    azul = PatternFill("solid", fgColor="0D4F7C")
    azul_claro = PatternFill("solid", fgColor="E8F4FD")
    verde = PatternFill("solid", fgColor="1E8449")
    gris = PatternFill("solid", fgColor="F2F2F2")
    borde_fino = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def encabezado(ws, columnas, color_fill=azul):
        ws.append(columnas)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = color_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = borde_fino
        ws.row_dimensions[1].height = 20

    def autoajustar(ws):
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    def estilo_fila(ws, fila_num, fill=None):
        for cell in ws[fila_num]:
            cell.border = borde_fino
            cell.alignment = Alignment(vertical='center')
            if fill:
                cell.fill = fill

    # 1. Socios
    ws = wb.create_sheet("Socios")
    encabezado(ws, ['ID', 'Nombre Completo', 'CI', 'Código Cliente', 'Teléfono', 'Estado', 'Observación Retiro', 'Fecha Retiro'])
    for i, s in enumerate(Socio.objects.all().order_by('nombre_completo'), 2):
        ws.append([
            str(s.pk), s.nombre_completo, s.ci, s.codigo_cliente or '',
            s.telefono or '', s.estado, s.observacion_retiro or '',
            s.fecha_retiro.strftime('%d/%m/%Y') if s.fecha_retiro else ''
        ])
        estilo_fila(ws, i, gris if i % 2 == 0 else None)
    autoajustar(ws)

    # 2. Medidores
    ws = wb.create_sheet("Medidores")
    encabezado(ws, ['ID', 'Número Medidor', 'Titular (Nombre)', 'Titular (CI)', 'Manzano', 'Parcela', 'Estado'])
    for i, m in enumerate(Medidor.objects.select_related('socio').all(), 2):
        ws.append([
            str(m.pk), m.numero_medidor or '', m.socio.nombre_completo,
            m.socio.ci, m.manzano or '', m.parcela or '', m.estado
        ])
        estilo_fila(ws, i, gris if i % 2 == 0 else None)
    autoajustar(ws)

    # 3. Tarifas
    ws = wb.create_sheet("Tarifas")
    encabezado(ws, ['ID', 'Nombre', 'Costo por m³', 'Cuota Fija', 'Multa Atraso', 'Reconexión', 'Limpieza Tanque', 'Instalación Clandestina', 'Multa Alteración', 'Falta Asamblea', 'Costo Afiliación', 'Días Gracia', 'Activa'])
    for i, t in enumerate(Tarifa.objects.all().order_by('-id_tarifa'), 2):
        ws.append([
            t.id_tarifa, t.nombre, float(t.costo_por_cubo), float(t.cuota_fija),
            float(t.multa_atraso), float(t.costo_reconexion), float(t.costo_limpieza_tanque),
            float(t.costo_instalacion_clandestina), float(t.costo_multa_alteracion),
            float(t.costo_falta_asamblea), float(t.costo_afiliacion), t.dias_gracia,
            'Sí' if t.activa else 'No'
        ])
        estilo_fila(ws, i, gris if i % 2 == 0 else None)
    autoajustar(ws)

    # 4. Lecturas
    ws = wb.create_sheet("Lecturas")
    encabezado(ws, ['ID', 'Número Medidor', 'Titular (CI)', 'Periodo', 'Lectura Anterior', 'Lectura Actual', 'Consumo m³', 'Fecha Lectura', 'Observación'])
    for i, l in enumerate(Lectura.objects.select_related('medidor', 'medidor__socio').all().order_by('periodo'), 2):
        ws.append([
            str(l.pk), l.medidor.numero_medidor or '', l.medidor.socio.ci,
            l.periodo, float(l.lectura_anterior), float(l.lectura_actual),
            float(l.consumo_cubos), l.fecha_lectura.strftime('%d/%m/%Y %H:%M') if l.fecha_lectura else '',
            l.observacion or ''
        ])
        estilo_fila(ws, i, gris if i % 2 == 0 else None)
    autoajustar(ws)

    # 5. Cobros
    ws = wb.create_sheet("Cobros")
    encabezado(ws, ['N° Recibo', 'Socio', 'CI', 'Medidor', 'Periodo', 'Fecha Emisión', 'Importe Consumo', 'Recargo Atraso', 'Monto Total', 'Estado'])
    for i, c in enumerate(Cobro.objects.select_related('socio', 'lectura', 'lectura__medidor').all().order_by('numero_recibo'), 2):
        ws.append([
            c.numero_recibo, c.socio.nombre_completo, c.socio.ci,
            c.lectura.medidor.numero_medidor or '', c.lectura.periodo,
            c.fecha_emision.strftime('%d/%m/%Y') if c.fecha_emision else '',
            float(c.importe_consumo), float(c.recargo_falta_pago),
            float(c.monto_total), c.estado_pago
        ])
        estilo_fila(ws, i, gris if i % 2 == 0 else None)
    autoajustar(ws)

    # 6. Pagos
    ws = wb.create_sheet("Pagos")
    encabezado(ws, ['ID', 'N° Recibo', 'Socio (CI)', 'Fecha Pago', 'Monto Pagado', 'Método'], verde)
    for i, p in enumerate(Pago.objects.select_related('recibo', 'recibo__socio').all().order_by('fecha_pago'), 2):
        ws.append([
            str(p.pk), p.recibo.numero_recibo, p.recibo.socio.ci,
            p.fecha_pago.strftime('%d/%m/%Y %H:%M') if p.fecha_pago else '',
            float(p.monto_pagado), p.metodo_pago or ''
        ])
        estilo_fila(ws, i, gris if i % 2 == 0 else None)
    autoajustar(ws)

    # 7. Afiliaciones
    ws = wb.create_sheet("Afiliaciones")
    encabezado(ws, ['ID', 'Socio (CI)', 'Medidor', 'Tipo', 'Monto (Bs)', 'Fecha Pago'])
    for i, a in enumerate(Afiliacion.objects.select_related('socio', 'medidor').all(), 2):
        ws.append([
            str(a.pk), a.socio.ci, a.medidor.numero_medidor if a.medidor else '',
            a.tipo, float(a.monto), a.fecha_pago.strftime('%d/%m/%Y') if a.fecha_pago else ''
        ])
        estilo_fila(ws, i, gris if i % 2 == 0 else None)
    autoajustar(ws)

    # 8. Usuarios
    ws = wb.create_sheet("Usuarios")
    encabezado(ws, ['Username', 'Nombre Completo', 'CI', 'Teléfono', 'Rol', 'Activo'])
    UsuarioModel = get_user_model()
    for i, u in enumerate(UsuarioModel.objects.all().order_by('rol', 'username'), 2):
        ws.append([
            u.username, u.get_full_name() or '', u.ci or '',
            u.telefono or '', u.rol, 'Sí' if u.activo else 'No'
        ])
        estilo_fila(ws, i, gris if i % 2 == 0 else None)
    autoajustar(ws)

    # 9. HOJA DE RESUMEN EJECUTIVO
    ws_resumen = wb.create_sheet("Resumen")
    ws_resumen.sheet_view.showGridLines = False

    ws_resumen['A1'] = 'BACKUP DEL SISTEMA DE AGUA POTABLE'
    ws_resumen['A1'].font = Font(bold=True, size=14, color="0D4F7C")
    ws_resumen['A2'] = 'Junta Vecinal "Paraíso"'
    ws_resumen['A2'].font = Font(size=11, color="555555")
    ws_resumen['A3'] = f'Generado el: {timezone.now().strftime("%d/%m/%Y %H:%M:%S")}'
    ws_resumen['A3'].font = Font(size=10, color="888888")

    ws_resumen['A5'] = 'RESUMEN Y TOTALES GENERALES'
    ws_resumen['A5'].font = Font(bold=True, size=11, color="0D4F7C")

    datos_resumen = [
        ('Total Socios Activos', Socio.objects.filter(estado='ACTIVO').count()),
        ('Total Medidores Activos', Medidor.objects.filter(estado='Activo').count()),
        ('Total Lecturas Registradas', Lectura.objects.count()),
        ('Total Cobros Generados', Cobro.objects.count()),
        ('Cobros Cancelados', Cobro.objects.filter(estado_pago='Cancelado').count()),
        ('Cobros Pendientes / Vencidos', Cobro.objects.filter(estado_pago__in=['Pendiente', 'En Revision', 'Vencido']).count()),
        ('Total Recaudado (Bs)', float(Pago.objects.aggregate(t=Sum('monto_pagado'))['t'] or 0)),
        ('Deuda Total Pendiente (Bs)', float(Cobro.objects.filter(estado_pago__in=['Pendiente', 'En Revision', 'Vencido']).aggregate(t=Sum('monto_total'))['t'] or 0)),
        ('Total Recaudado por Afiliaciones (Bs)', float(Afiliacion.objects.aggregate(t=Sum('monto'))['t'] or 0)),
    ]

    for idx, (label, valor) in enumerate(datos_resumen, 7):
        ws_resumen[f'A{idx}'] = label
        ws_resumen[f'B{idx}'] = valor
        ws_resumen[f'A{idx}'].font = Font(size=10)
        ws_resumen[f'B{idx}'].font = Font(bold=True, size=10)
        ws_resumen[f'A{idx}'].border = borde_fino
        ws_resumen[f'B{idx}'].border = borde_fino
        if idx % 2 == 0:
            ws_resumen[f'A{idx}'].fill = azul_claro
            ws_resumen[f'B{idx}'].fill = azul_claro

    ws_resumen.column_dimensions['A'].width = 40
    ws_resumen.column_dimensions['B'].width = 25

    fecha_str = timezone.now().strftime('%Y%m%d_%H%M')
    nombre_archivo = f'backup_sistema_agua_{fecha_str}.xlsx'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response


@login_required
@es_admin_o_tesorero
def descargar_plantilla_excel(request):
    """Genera y descarga una plantilla vacía para la importación masiva de datos."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="0D4F7C")
    borde_fino = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def crear_hoja(nombre, columnas):
        ws = wb.create_sheet(nombre)
        ws.append(columnas)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = borde_fino
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = 25

    crear_hoja("Socios", ['ci', 'nombre_completo', 'codigo_cliente', 'telefono', 'estado'])
    crear_hoja("Medidores", ['numero_medidor', 'ci_socio', 'manzano', 'parcela', 'estado'])
    crear_hoja("Lecturas", ['numero_medidor', 'periodo', 'lectura_anterior', 'lectura_actual', 'observacion'])
    crear_hoja("Tarifas", ['nombre', 'costo_por_cubo', 'cuota_fija', 'multa_atraso', 'dias_gracia', 'activa'])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Plantilla_Importacion_Sistema_Agua.xlsx"'
    wb.save(response)
    return response


@login_required
@es_admin_o_tesorero
def importar_excel(request):
    """
    Importa masivamente datos desde un archivo Excel.
    Aplica Opción A (actualiza duplicados) y genera Cobros automáticos para lecturas.
    Requiere confirmación con contraseña.
    """
    if request.method != 'POST':
        return redirect('backup_vista')

    password_confirmacion = request.POST.get('password_confirmacion', '')
    if not request.user.check_password(password_confirmacion):
        messages.error(request, 'Contraseña incorrecta. Se canceló la importación.')
        return redirect('backup_vista')

    archivo = request.FILES.get('archivo_excel')
    if not archivo or not archivo.name.endswith(('.xlsx', '.xls')):
        messages.error(request, 'Por favor, selecciona un archivo de Excel válido (.xlsx).')
        return redirect('backup_vista')

    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
        socios_creados = medidores_creados = lecturas_creadas = tarifas_creadas = 0

        with transaction.atomic():
            # 1. Cargar Socios
            if "Socios" in wb.sheetnames:
                ws = wb["Socios"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0] or not row[1]:
                        continue
                    ci, nombre = str(row[0]).strip(), str(row[1]).strip()
                    codigo = str(row[2]).strip() if len(row) > 2 and row[2] else None
                    telef = str(row[3]).strip() if len(row) > 3 and row[3] else None
                    est = str(row[4]).strip() if len(row) > 4 and row[4] else 'ACTIVO'

                    Socio.objects.update_or_create(
                        ci=ci,
                        defaults={
                            'nombre_completo': nombre,
                            'codigo_cliente': codigo,
                            'telefono': telef,
                            'estado': est,
                        }
                    )
                    socios_creados += 1

            # 2. Cargar Medidores
            if "Medidores" in wb.sheetnames:
                ws = wb["Medidores"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0] or not row[1]:
                        continue
                    num_med, ci_soc = str(row[0]).strip(), str(row[1]).strip()
                    manzano = str(row[2]).strip() if len(row) > 2 and row[2] else None
                    parcela = str(row[3]).strip() if len(row) > 3 and row[3] else None
                    estado = str(row[4]).strip() if len(row) > 4 and row[4] else 'Activo'

                    socio = Socio.objects.filter(ci=ci_soc).first()
                    if socio:
                        Medidor.objects.update_or_create(
                            numero_medidor=num_med,
                            defaults={
                                'socio': socio,
                                'manzano': manzano,
                                'parcela': parcela,
                                'estado': estado,
                            }
                        )
                        medidores_creados += 1

            # 3. Cargar Tarifas
            if "Tarifas" in wb.sheetnames:
                ws = wb["Tarifas"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue
                    nom = str(row[0]).strip()
                    costo_cubo = Decimal(str(row[1] or '0'))
                    cuota_f = Decimal(str(row[2] or '0'))
                    multa_a = Decimal(str(row[3] or '0'))
                    dias_g = int(row[4] or 0)
                    activa = str(row[5]).strip().lower() in ['si', 'sí', 'true', '1'] if len(row) > 5 else False

                    if activa:
                        Tarifa.objects.update(activa=False)

                    Tarifa.objects.update_or_create(
                        nombre=nom,
                        defaults={
                            'costo_por_cubo': costo_cubo,
                            'cuota_fija': cuota_f,
                            'multa_atraso': multa_a,
                            'dias_gracia': dias_g,
                            'activa': activa,
                        }
                    )
                    tarifas_creadas += 1

            # 4. Cargar Lecturas (Gatilla el post_save para generar Cobros automáticamente)
            if "Lecturas" in wb.sheetnames:
                ws = wb["Lecturas"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0] or not row[1]:
                        continue
                    num_med, per = str(row[0]).strip(), str(row[1]).strip()
                    l_ant = Decimal(str(row[2] or '0'))
                    l_act = Decimal(str(row[3] or '0'))
                    obs = str(row[4]).strip() if len(row) > 4 and row[4] else None

                    medidor = Medidor.objects.filter(numero_medidor=num_med).first()
                    if medidor:
                        lectura, created = Lectura.objects.update_or_create(
                            medidor=medidor,
                            periodo=per,
                            defaults={
                                'lectura_anterior': l_ant,
                                'lectura_actual': l_act,
                                'observacion': obs,
                                'creado_por': request.user,
                            }
                        )
                        lecturas_creadas += 1

        messages.success(
            request,
            f'¡Importación completada exitosamente! Registros procesados: '
            f'{socios_creados} socios, {medidores_creados} medidores, {tarifas_creadas} tarifas, '
            f'{lecturas_creadas} lecturas con sus cobros automáticos.'
        )

    except Exception as e:
        messages.error(request, f'Ocurrió un error al procesar el archivo Excel: {str(e)}')

    return redirect('backup_vista')